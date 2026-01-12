"""Excel styling helpers for generated workbooks."""

from __future__ import annotations

from openpyxl import load_workbook
from openpyxl.styles import NamedStyle


def _has_named_style(workbook, style_name: str) -> bool:
    """Return True if the workbook already defines the named style."""
    for named_style in workbook.named_styles:
        if getattr(named_style, "name", named_style) == style_name:
            return True
    return False


def _ensure_named_style(workbook, style: NamedStyle) -> None:
    """Ensure a NamedStyle exists on the workbook."""
    if not _has_named_style(workbook, style.name):
        workbook.add_named_style(style)


def apply_styles(file_name: str, date_column: int, montant_column: int) -> None:
    """Apply date and amount styles to the Excel file.

    Args:
        file_name: Path to the Excel file to modify.
        date_column: 1-based index of the date column.
        montant_column: 1-based index of the amount column.
    """
    wb = load_workbook(file_name)
    ws = wb.active

    date_style = NamedStyle(name="date_style", number_format="DD-MM-YY")
    montant_style = NamedStyle(
        name="montant_style",
        number_format="#,##0.00 €;[RED]- #,##0.00 €",
    )
    _ensure_named_style(wb, date_style)
    _ensure_named_style(wb, montant_style)

    for row in ws.iter_rows(min_row=2, min_col=date_column, max_col=date_column):
        for cell in row:
            cell.style = date_style.name

    for row in ws.iter_rows(min_row=2, min_col=montant_column, max_col=montant_column):
        for cell in row:
            cell.style = montant_style.name

    wb.save(file_name)
    print(f"Styles appliqués et fichier sauvegardé : {file_name}")
